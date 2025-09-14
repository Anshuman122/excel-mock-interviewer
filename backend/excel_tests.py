import openpyxl

def validate_excel(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    report = {"score": 0, "pass_fail": False, "notes": []}

    # Example 1: Check sheet exists
    if "Sheet1" in wb.sheetnames:
        report["score"] += 3
    else:
        report["notes"].append("Sheet1 missing")

    # Example 2: Check formula in cell
    ws = wb["Sheet1"]
    cell = ws["B10"]
    if cell.value and "SUM(A2:A10)" in str(cell.value):
        report["score"] += 7
    else:
        report["notes"].append("SUM formula not found in B10")

    # Pass if score >= 8
    report["pass_fail"] = report["score"] >= 8
    return report
